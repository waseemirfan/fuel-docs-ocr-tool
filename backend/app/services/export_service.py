import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models.document import Document, ExtractionField

COLUMNS = ["ID", "Filename", "Date", "Manifest No", "BoL", "Delivery Point",
           "Regular", "Super", "Diesel", "Overall Confidence %", "Status", "Uploaded At"]

FIELD_MAP = {
    "Date": "date",
    "Manifest No": "manifest_no",
    "BoL": "bol",
    "Delivery Point": "delivery_point",
    "Regular": "regular",
    "Super": "super",
    "Diesel": "diesel",
}

def _effective_value(field: ExtractionField) -> str:
    return field.corrected_value if field.corrected_value is not None else (field.extracted_value or "")


def generate_excel(documents: list[Document]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Extractions"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    ws.row_dimensions[1].height = 30

    for row_idx, doc in enumerate(documents, start=2):
        field_lookup = {f.field_name: f for f in doc.fields}
        confidence_pct = f"{doc.overall_confidence:.1f}%" if doc.overall_confidence is not None else ""

        row_data = [
            doc.id,
            doc.filename,
        ]
        for col_label in ["Date", "Manifest No", "BoL", "Delivery Point", "Regular", "Super", "Diesel"]:
            fn = FIELD_MAP[col_label]
            field = field_lookup.get(fn)
            row_data.append(_effective_value(field) if field else "")

        row_data += [confidence_pct, doc.status.value, doc.uploaded_at.strftime("%Y-%m-%d %H:%M")]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F2F7FB")

    for col_idx in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
